#!/usr/bin/env python3
"""
Script para generar la plantilla Excel de importación de clientes
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Crear DataFrame con los encabezados
headers = [
    'TipoCliente*', 'Nombre*', 'DNI/CIF', 'DNI Representante', 'Email', 
    'Teléfono', 'Dirección', 'Número', 'Escalera', 'Piso', 
    'Puerta', 'Aclarador', 'Población', 'Provincia', 'Código Postal', 
    'IBAN', 'Representante', 'Comercial', 'Observaciones'
]

# Datos de ejemplo
ejemplo = {
    'TipoCliente*': 'Particular',
    'Nombre*': 'Ejemplo Cliente S.L.',
    'DNI/CIF': 'B12345678',
    'DNI Representante': '12345678A',
    'Email': 'cliente@ejemplo.com',
    'Teléfono': '912345678',
    'Dirección': 'Calle Mayor',
    'Número': '10',
    'Escalera': 'A',
    'Piso': '3',
    'Puerta': 'B',
    'Aclarador': 'Junto al parque',
    'Población': 'Madrid',
    'Provincia': 'Madrid',
    'Código Postal': '28001',
    'IBAN': 'ES1234567890123456789012',
    'Representante': 'Juan Pérez',
    'Comercial': 'María García',
    'Observaciones': 'Cliente potencial importante'
}

# Crear hoja de clientes con ejemplo
df_clientes = pd.DataFrame([ejemplo])

# Crear hoja de instrucciones
instrucciones_texto = [
    ['INSTRUCCIONES PARA IMPORTAR CLIENTES'],
    [''],
    ['Campos obligatorios (marcados con *):'],
    ['- TipoCliente: Debe ser exactamente "Particular" o "Empresa"'],
    ['- Nombre: Nombre o razón social del cliente'],
    [''],
    ['Campos opcionales:'],
    ['- DNI/CIF: Documento de identificación'],
    ['- DNI Representante: DNI del representante legal'],
    ['- Email: Correo electrónico (se valida el formato)'],
    ['- Teléfono: Número de teléfono de contacto'],
    ['- Dirección, Número, Escalera, Piso, Puerta: Dirección completa del cliente'],
    ['- Aclarador: Información adicional de ubicación (ej: "Junto al parque")'],
    ['- Población, Provincia, Código Postal: Datos de localización'],
    ['- IBAN: Cuenta bancaria (máximo 34 caracteres)'],
    ['- Representante: Nombre del representante legal'],
    ['- Comercial: Nombre del comercial asignado al cliente'],
    ['- Observaciones: Notas adicionales sobre el cliente'],
    [''],
    ['Notas importantes:'],
    ['- NO elimine la primera fila de encabezados en la hoja "Clientes"'],
    ['- Puede eliminar la fila de ejemplo antes de importar sus datos'],
    ['- Añada tantas filas como clientes necesite importar'],
    ['- Los campos vacíos se guardarán como NULL en la base de datos'],
    ['- La fecha de alta se asignará automáticamente al momento de importar'],
    ['- El formato del archivo debe ser .xlsx'],
    [''],
    ['Validaciones que realiza el script:'],
    ['- TipoCliente debe ser "Particular" o "Empresa" (exactamente, respetando mayúsculas)'],
    ['- Nombre es obligatorio y no puede estar vacío'],
    ['- Email debe tener un formato válido si se proporciona'],
    ['- IBAN no puede exceder 34 caracteres'],
    ['- Todos los campos de texto respetan los límites de la base de datos'],
    [''],
    ['Ejemplos de valores válidos:'],
    ['TipoCliente: "Particular" o "Empresa"'],
    ['Nombre: "Juan Pérez García" o "Empresa Ejemplo S.L."'],
    ['Email: "ejemplo@dominio.com"'],
    ['IBAN: "ES1234567890123456789012"'],
    [''],
    ['Para más información, consulte el archivo IMPORTACION_CLIENTES.md']
]

df_instrucciones = pd.DataFrame(instrucciones_texto)

# Guardar en Excel
archivo = 'plantilla_clientes.xlsx'
with pd.ExcelWriter(archivo, engine='openpyxl') as writer:
    df_clientes.to_excel(writer, sheet_name='Clientes', index=False)
    df_instrucciones.to_excel(writer, sheet_name='Instrucciones', index=False, header=False)

# Aplicar formato
wb = load_workbook(archivo)

# Formatear hoja de Clientes
ws_clientes = wb['Clientes']
header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
header_font = Font(bold=True)

for cell in ws_clientes[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Ajustar ancho de columnas
for column in ws_clientes.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws_clientes.column_dimensions[column_letter].width = adjusted_width

# Formatear hoja de Instrucciones
ws_instrucciones = wb['Instrucciones']
ws_instrucciones['A1'].font = Font(bold=True, size=14)
ws_instrucciones.column_dimensions['A'].width = 100

# Guardar
wb.save(archivo)
print(f"✓ Plantilla creada exitosamente: {archivo}")
print(f"✓ Incluye hoja 'Clientes' con ejemplo y hoja 'Instrucciones' con ayuda detallada")
